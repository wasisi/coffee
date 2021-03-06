import openpyxl
import datetime
import os
import csv

NoneString=""

#the columns of a row that we are interested in
#These correspond to excel sheet column letters
#
#Field-Column map
#A=Transanr,        B=Lotnr,        C=Marks,
#D=Grade,           E=Bags,         F=Weight,
#G=Saleno,          H=BagsBought    I=WeightBought
#J=BuyerCode,       K=Price,        L=SeatNr,
#M=AUTCODE,         N=Status,       O=Datum,        P=TijD

#column names over which the script loops
#to extract data
interestCols='ABCDEFGHIJKLMNO'

#map for the months
months={'01':'Jan','02':'Feb','03':'Mar',
        '04':'Apr','05':'May','06':'June',
        '07':'July','08':'Aug','09':'Sept',
        '10':'Oct','11':'Nov','12':'Dec'}




#map for the error codes with explanation
error_map={"01":"Missing TRANSANR column value",
           "02":"Missing LOTNR column value",
           "03":"Missing MARKS column value",
           "031":"Missing Reference In Marks Column",
           "04":"Missing GRADE column value",
           "05":"Missing BAGS column value",
           "06":"Missing WEIGHT column value",
           "07":"Missing SALENO column value",
           "08":"Missing BAGSBOUGHT column value",
           "09":"Missing WGTBOUGHT column value",
           "10":"Missing BUYERCODE column value",
           "11":"Missing PRICE column value",
           "12":"Missing SEATNR column value",
           "13":"Missing AUCTCODE column value",
           "14":"Missing STATUS column value",
           "15":"Missing DATUM column value",
           "16":"Missing TIJD column value",
           "17":"Couldn't write value column price or weightBought are invalid"}

#maps the column with the error code
column_error_map={"A":"01",
                  "B":"02",
                  "C":"03",
                  "D":"04",
                   "E":"05",
                   "F":"06",
                   "G":"07",
                   "H":"08",
                   "I":"09",
                   "J":"10",
                   "K":"11",
                   "L":"12",
                   "M":"13",
                   "N":"14",
                   "O":"15",
                   "P":"16"
                }

def correct_mark_format(cell_value):

	if cell_value==None:
		return NoneString

	#begin with removing the spaces in the given
	#string. Strings are immutables in Python
	marks2 = cell_value.replace(" ","")

	# the string using / as delimeter 
	separate_list = marks2.split('/')

	#no factory number and I don't
	#know what to do. I am simply returning
  #the error code
	if len(separate_list) <= 2:
		return "031"

	#get the factory number
	factory_num = separate_list[2]

	#find and replace the O to zero string
	factory_num.replace('O','0')

	separate_list[2]=factory_num

	marks2=""
	#merge the list and return the value
	for i in range(len(separate_list)):
		marks2 += separate_list[i]+'/'

	#remove full stops
	marks2 = marks2.replace('.',"")
	marks2_split = marks2.split('/')
	return (cell_value,marks2,marks2_split[2],marks2_split[2],marks2_split[0])


def process_datum(cell_value):
    """
    Proceccess the Datum column to produce an ISODate
    and Season
    """
    if cell_value==None or cell_value=="":
		    return "15"

    value = str(cell_value.date())
    value_list = value.split('-')
    year =  value_list[0]
    month = value_list[1]
    day = value_list[2]
    yearint = int(year)

    #which month this is
    if month == '10' or month == '11' or month == '12':
        #season starts in October so it is this year plus 1
        yearnext = yearint+1
        return (str(year)+'-'+str(month)+'-'+str(day),str(year)+'-'+str(yearnext))
    else:
        #then it is this year minus 1
        yearprev = yearint-1
        return (str(year)+'-'+str(month)+'-'+str(day),str(yearprev)+'-'+str(year))

def correct_output_csv_file(csv_file_name):
    """
    Sets the name of the output CSV file.
    TODO: Maybe we would like a regular expression
     for this function
    """
    if csv_file_name =="" or csv_file_name==None:
        now = datetime.datetime.now()
        csv_file_name = 'output-'+str(now)+'.csv'
        csv_file_name = csv_file_name.replace(":","-")

    #replace any blanks
    csv_file_name = csv_file_name.replace(" ","_")

    #try to find the pattern .csv at the end of the string
    # if it does not exist append it at the end
    filename,file_extension = os.path.splitext(csv_file_name)

    if file_extension == '.csv':
        return csv_file_name

    #otherwise simply keep the file name
    #and add the extension
    return filename+'.csv'


def write_error_output(filename,errors):
    """
    Write a file describing the row errors found
    """
    try:
        csv_file = open(filename,'wt')
        csvwriter = csv.writer(csv_file,lineterminator='\n')

        for error in errors:
            csvwriter.writerow(error)
    finally:
        csv_file.close()


def cleanup(excel_in_filename,csv_out_filename):
    """
    Main function to call for cleanup
    """
    try:
        #the excel doc to work om
        workbook = openpyxl.load_workbook(excel_in_filename)

        csv_file = open(csv_out_filename,'wt')

        #the csv writer to write into
        csvwriter = csv.writer(csv_file,lineterminator='\n')

        #the sheet name in the workbook
        sheets = workbook.get_sheet_names()

        #how many sheets the document has
        nsheets = len(sheets)

        #safe guard in the case no sheets exist
        if nsheets == 0:
            csvwriter.writerow(("No sheets found in "+excel_in_filename,))
            return

        #by default we only process the zero sheet
        #perhaps we can change this; add some color in the future
        if nsheets > 1:
            print("===================================================================")
            print("WARNING: More than one sheets found. Processing only the first one.")
            print("===================================================================")

        #get the sheet. Assume that the workbook has
        #only one sheet? Validate?
        sheet = workbook.get_sheet_by_name(sheets[0])

        csvwriter.writerow(('#TRANSNR','LOTNT','MARKS', 'MARKS2','REF','REF2', 'BAGMARK','GRADE-GR',
                            'BAGSNR','WEIGHT-Kgr','SALENO','BAGSBOUGHTNR','WEIGHTBOUGHT-Kgr',
                            'BUYERCODE','PRICE','SEATNR','AUCTCODE','STATUS','ISODATE', 'SEASON','VALUE'))

        print("\tProcessing %d rows: "%sheet.max_row)

        #array holding the failed rows
        failed_rows=[]

        #an array that indicates if a row
        #failed or not
        failed_row=[False for x in range(0,sheet.max_row)]

        for row in range(2,sheet.max_row):
            row_vals=[]
            weightBought=None
            price = None
            for col in interestCols:
                cell = "{}{}".format(col,row)

                if col == 'C': #this is the marks column
                    cell_value = correct_mark_format(sheet[cell].value)
                    if cell_value == "031":
                        failed_rows.append((row,error_map["031"]))
                        failed_row[row]=True
                    else:
                        for val in cell_value:
                            row_vals.append(val)
                elif col=='O':
                    values = process_datum(sheet[cell].value)
                    if values == column_error_map[col]:
                        failed_rows.append((row,error_map[column_error_map[col]]))
                        failed_row[row]=True
                    else:
                        for val in values:
                            row_vals.append(val)
                elif col=='I':

                    if sheet[cell].value == None or sheet[cell].value == "":
                        failed_rows.append((row,error_map[column_error_map[col]]))
                        failed_row[row]=True
                    else:
                        #cache the weight bought for value calculation
                        weightBought = float(sheet[cell].value)
                        row_vals.append(sheet[cell].value)
                elif col=='K':
                    if sheet[cell].value == None or sheet[cell].value == "":
                        failed_rows.append((row,error_map[column_error_map[col]]))
                        failed_row[row]=True
                    else:
                        #cache the price for value calculation
                        price = float(sheet[cell].value)
                        row_vals.append(sheet[cell].value)
                else:
                    row_vals.append(sheet[cell].value)

            if price != None and weightBought != None:
                #we processed the sheet let's calculate the value
                value = (weightBought/50.)*price
                row_vals.append(value)
            else:
                failed_rows.append((row,error_map["17"]))
                failed_row[row]=True

            #let's write the row in the specified
            #we only write if the row has not failed
            #in one way or another
            if failed_row[row]==False:
                #make a tuple from the list values
                row_vals = tuple(row_vals)
                csvwriter.writerow(row_vals)

        return failed_rows
    except FileNotFoundError:
        print("File "+excel_in_filename+" does not exist")
    finally:
        csv_file.close()


if __name__ == '__main__':
    print("=======================================================================")
    print("\tStart Processing")
    print("=======================================================================")
    excel_doc = input("\tName of excel document or path to the document: ")

    #make sure that something is given
    if excel_doc == "":
        raise ResourceWarning("\tEntered empty excel file name for processing. Exiting")

    print("\tExcel file given %s"%excel_doc)

    #the name of the output of the CSV file
    csv_file_name = input("\tEnter CSV file name. Default is ouput-yyyy-mm-dd_hh-mm-ss.csv. Leave blank if default is ok: ")
    csv_file_name = correct_output_csv_file(csv_file_name)
    print("\tSaving at file: ",csv_file_name)

    #clean up the given excel and output to csv
    failed_rows = cleanup(excel_doc,csv_file_name)
    print("====================ERROR REPORT=====================================")
    if not failed_rows:
        print("\tNo failed rows detected")

    else:
        print("\tFailed rows occured. Number of failed rows: %d "%len(failed_rows))
        errfilename = 'error-'+csv_file_name
        print("\tOutputing error file at %s "%errfilename)
        write_error_output(errfilename,failed_rows)
    print("=======================================================================")
    print("\tEnd Processing")
    print("=======================================================================")
	
